#!/Users/jayhan/VirtualEnvs/flask_py3/bin python
# -*- coding: utf-8 -*-
import xlrd
import openpyxl
import numpy as np
from functools import reduce
import datetime
import os
import re
from py2neo import Graph,Node,Relationship
strDTime_default='17 NOV 01  08:30:00'
#unicode = str ### set during python3 enviornment
class DTime(object):

    def __init__(self,strDTime=strDTime_default):
        self._datetime = datetime.datetime.strptime(strDTime,"%y %b %d  %X")
        self.year = self._datetime.year
        self.month = self._datetime.month
        self.day = self._datetime.day
        self.hour = self._datetime.hour
        self.minute = self._datetime.minute
        self.second = self._datetime.second

class Table:
    def __init__(self, arr=[], col=[],dt=DTime()):
        self.array = arr
        self.columns = col
        self.strDateTime = dt

class Jsw:
    '''
    info=[connector1,pin1,connector2,pin2,chapter,testType]
    '''
    columns = ['connector1','pin1','connector2','pin2','chapter','pin1Type','pin2Type']
    test_type = ['continuity','insulation']
    sheet_in = [u'连续性测试表', u'接地线导通测试表']
    pin_type = ['auto','tb','nap']

    def __init__(self):
        self.nRows = 0
        self.nConnection=0
        self.tables =[]

    def open_excel(self, fileName='file.xls'):
        data = xlrd.open_workbook(fileName)
        return data
        #try:
         #   data = xlrd.open_workbook(fileName)
         #   return data
       # except Exception e:
        #    print str(e)

    def excel_table_byindex(self,fileName='file.xls',colnameindex=0,by_index=0):
        data = self.open_excel(fileName)
        table = data.sheets()[by_index]
        nrows = table.nrows-1 #行数
        cols = table.ncols #列数
        colnames =  table.row_values(colnameindex) #某一行数据
        list = []
        for rownum in range(1,nrows):
             row = table.row_values(rownum)
             if row:
                list.append(row)
        return list

    #根据名称获取Excel表格中的数据
    # 参数:file：Excel文件路径，colnameindex：表头列名所在行，by_name：Sheet1名称
    def excel_table_byname(self,fileName= 'file.xls',colnameindex=0,by_name=u'Sheet1'):
        data = self.open_excel(fileName)
        table = data.sheet_by_name(by_name)
        self.nRows += table.nrows-1#行数
        #print(self.nrows, self.ncols)
        self.colnames = table.row_values(colnameindex) #某一行数据
        list =[]
        for rownum in range(1,table.nrows):
             row = table.row_values(rownum)
             if row:
                list.append(row)
        return list

    def process(self,path):
        data_in = self.excel_table_byname(fileName=path, by_name=Jsw.sheet_in[0])
        #print(data_in)
        data_in = self._strcleanning(data_in)
        #print(data_in)
        self.tables.append(Table(self._pinType(data_in), Jsw.columns))

        data_in2 = self.excel_table_byname(fileName=path, by_name=Jsw.sheet_in[1])
        data_in2 = self._strcleanning(data_in2)
        self.tables.append(Table(self._pinType(data_in2),Jsw.columns))

    def _strcleanning(self,lstTable):
        '''
        get useful information from pd at column of [1, 2, 4, 5, 7]
        :param pd:
        :return:
        '''

        lstTable_2 = []
        colSel = [1, 2, 4, 5, 7]
        self.colnames = [unicode(self.colnames[i]).replace(' ','') for i in colSel]
        for row in lstTable:
            #print(row)
            row = [unicode(row[i]).replace(' ','') for i in colSel]
            #print(row)
            lstTable_2.append(row)
        ## select connectorName1, pinIndex1, connectorName2, pinIndex2, chapterName from DataFrame
        return lstTable_2

    def _hasTB(self, row_data):
        '''
        whether 'TB' in row elements.upper()
        :param row_data:
        :return:
        '''
        for data in row_data:
            if u"TB" in data.upper():
                return True
        return False

    def _valid(self, cell):
        '''
        cnt1,cnt2 must be a combination of 0-9,a-z,A-Z,'-','*'
        '''
        mat1 = re.match("[\w-]+", cell)
        if mat1 and cell == mat1.group():
            return True
        return False


    def _pinType(self,array):
        row = len(array)
        pin1Type = None
        pin2Type = None
        for r in range(row):
            # print(r)
            cnt1, index1, cnt2, index2, chapter = array[r]
            pin1Type='auto'
            pin2Type='auto'
            if self._hasTB((cnt1,index1)):
                pin1Type = 'tb'
            elif not self._valid(cnt1):
                pin1Type = 'nap'
            if self._hasTB((cnt2,index2)):
                pin2Type = 'tb'
            elif not self._valid(cnt2):
                pin2Type = 'nap'
            array[r].append(pin1Type)
            array[r].append(pin2Type)
        return array


class Pvg:
    '''
    info=[connector1,pin1,connector2,pin2,chapter,\
    testType,status,value,unit,addr1,addr2]
    '''

    columns = ["pin1", "pin2", "testType", "status", "value", "unit", "pin1_addr", "pin2_addr"]
    typedic = {'FC': 'insulation', 'CC': 'continuity'}
    info_pattern = r"(?<=:)\s+([A-Z]{2})\s+([0-9]+)\s+([\S]+)\s*:\s+([0-9]+)\s+([A-Z]+)\s+([<>0-9.MK]+)\s+([A-Z]+)\s+([\S]+)"
    time_pattern = r"(?<=测试中止)\s+([0-9]+[\s\S]+[0-9]+)\s*分析仪停止"

    def __init__(self):
        '''
        '''
        self.tables = []
        self.nRows = 0
        self.nConnection = 0

    def process(self,path):
        '''
        In -- TXT file of testing report
        Return -- [[ ],...]
        '''
        table = Table()
        fp = open(path, U'r')
        txt = fp.read()
        lists = []
        time = strDTime_default
        re1 = re.compile(Pvg.info_pattern+"|"+Pvg.time_pattern)
        for mat in re1.finditer(txt):
            # print(mat.groups())
            line = mat.groups() #["command","addr1","pin1","addr2","status","value","unit","pin2"]
            if reduce(lambda x,y: x and y,line[:8]):
                # print("line[:8]",line[:8])
                pin1,pin2,pin1_addr, pin2_addr, status, value, unit = line[2],line[7],line[1], line[3], line[4], line[5], line[6]
                test_type = Pvg.typedic.get(line[0], 'NULL')
                line = [pin1, pin2, test_type, status, value, unit, pin1_addr, pin2_addr]
                lists.append(line)
                self.nRows += 1
                self.nConnection += 1
            elif line[8]:
                time = line[8]

        fp.close()
        table.array = lists#pd.DataFrame(lists, columns=Pgv.columns)
        table.columns = Pvg.columns
        table.dateTime = DTime(time)
        self.tables.append(table)
        # print(self.pdTestLists.shape)


class Neo4j(object):
    try:
        _graph = Graph('bolt://127.0.0.1:7687',user="neo4j",password='comac.123')
    except:
        print("Please check database Neo4j!")
        exit(-1);

    _node_lables = ["pin"]
    _rel_labeles = ["continuity","insulation"]
    _jsw_columns = ['connector1','pin1','connector2','pin2','chapter','pin1Type','pin2Type']
    _pgv_columns = [u'connector1',u'pin1',u'connector2',u'pin2',u'testType',u'status',u'value',u'unit',u'addr1',u'addr2']
    def __init__(self,pace=1):
        self.times = 0;
        self.pace = pace;

    def jsw_upload(self,objJsw,sequence_offset=0):
        '''
        upload new node and relationship to data base
        :param info: DataFrame with column name = _jsw_columns
        :return: True/False
        '''
        nTable,nTable_total = 0,len(objJsw.tables)
        #print(nTable,nTable_total)
        r = sequence_offset
        for table in objJsw.tables:
            colName = table.columns
            array = table.array
            if len(colName) != len(Neo4j._jsw_columns):
                return False
            if not reduce(lambda x, y: x and y, [colName[i] == Neo4j._jsw_columns[i] for i in range(len(colName))]):
                print("improper data format, please check!")
                return False
            # try:
            #     # self.clear()
            #     pass
            #     #print("to be")
            #     # set constraint
            # except:
            #     print("Please check connection of Neo4j Database!")
            #     return False
            node_gnd = Node(Neo4j._node_lables[0], connectorName='GND', pinIndex='', fullName='GND', pinType='auto')
            nRow,nRow_total= 0,len(array)
            #print(array)
            for row in array:
                self.times += 1
                if(self.times%self.pace == 0):
                    print("progress: %d%%"%(self.times//self.pace))

                # print(row)
                cntName1, pin1, cntName2, pin2, chapter_details, pin1Type, pin2Type = row
                chapter = chapter_details[:2]
                label1,label2 = Neo4j._rel_labeles
                if pin1 is np.nan or not pin1:
                   fullName1 = unicode(cntName1)
                else:
                   fullName1 = unicode(cntName1)+'-'+unicode(pin1)
                if pin2 is np.nan or not pin2:
                   fullName2 = unicode(cntName2)
                else:
                   fullName2 = unicode(cntName2)+'-'+unicode(pin2)
                node1 =Node(Neo4j._node_lables[0],connectorName=cntName1, pinIndex=pin1,fullName=fullName1,pinType=pin1Type)
                node2 =Node(Neo4j._node_lables[0],connectorName=cntName2, pinIndex=pin2,fullName=fullName2,pinType=pin2Type)
                Neo4j._graph.merge(node1)
                Neo4j._graph.merge(node2)
                if nTable == 0:
                    rel1 = Relationship(node1, label1, node2, chapter=chapter, chapter_details=chapter_details,status='NULL',times=0,sequence= r)
                    Neo4j._graph.merge(node_gnd)
                    Neo4j._graph.merge(rel1)
                    rel2 = Relationship(node1, label2, node_gnd, chapter=chapter, chapter_details=chapter_details,status='NULL', times=0, sequence=nRow_total+r)
                    Neo4j._graph.merge(rel2)
                elif nTable == 1:
                    rel1 = Relationship(node1, label1, node2, chapter=chapter,chapter_details=chapter_details,status='NULL', times=0, sequence=r)
                    Neo4j._graph.merge(rel1)
                nRow += 1
                r += 1
            nTable += 1
        return r

    def pvg_update(self,objPvg):
        '''
        Only modify property of existing relationship
        :param info:DataFrame with column name = _pvg_columns
        :return:False/True
        '''
        nTable, nTable_total = 0, len(objPvg.tables)
        high_count = 0

        for table in objPvg.tables:
            nRow, nRow_total = 0, len(table.array)
            colName = table.columns
            array = table.array

            # if len(colName) != len(Neo4j._pgv_columns):
            #     print("columns length not equal, please check!")
            #     return False
            # if not reduce(lambda x, y: x and y, [colName[i] == Neo4j._jsw_columns[i] for i in range(len(colName))]):
            #     print("improper data format, please check!")
            #     return False

            for row in array:

                fullName1,fullName2,testType,status,value,unit,addr1,addr2 = row
                if status == "HIGH":
                    high_count += 1
                # print("{0}/{1}".format(nRow + 1, nRow_total))

                query='''
                MATCH (pin1:pin)-[rel]->(pin2:pin)
                WHERE pin1.fullName = {name1} and pin2.fullName = {name2}
                WITH rel,pin1,pin2,rel.times as t
                SET rel.times = t+1,rel.status = {status},rel.value = {value},rel.unit = {unit},pin1.addr= {addr1},pin2.addr = {addr2}
                RETURN pin1,rel,pin2
                '''
                data = Neo4j._graph.run(query,name1=fullName1,name2=fullName2,\
                                     status=status,value=value,unit=unit,addr1=addr1,addr2=addr2)
                if not data:
                    print("NOT FOUND:",fullName1,fullName2,status,value,unit,addr1,addr2)
#                    return False
                nRow += 1
                self.times += 1
                if (self.times % self.pace == 0):
                    print("progress: %d%%" % (self.times // self.pace))
        print("high count: %d"%high_count)
        return True

    def clear(self):
        Neo4j._graph.delete_all()


    def prog(self, chapter=None,label=None):
        '''
        :label: label type of relationship,eg. continuity,insulation
        :return: (json objects)
        '''
        label_category=['continuity','insulation']

        if chapter and label in label_category:
            query = '''
            MATCH (pin1:pin)-[rel:{label}]->(pin2:pin)
            WHERE rel.status='HIGH' and rel.chapter='{chapter}'
            RETURN type(rel) as testType, pin1.fullName AS PIN1,pin2.fullName AS PIN2,rel.chapter as CHAPTER
            ORDER BY rel.sequence
            '''
            query = query.format(label=label,chapter=chapter)

        elif not chapter and label in label_category:
            query = '''
            MATCH (pin1:pin)-[rel:{label}]->(pin2:pin)
            WHERE rel.status='HIGH'
            RETURN type(rel) as testType, pin1.fullName AS PIN1,pin2.fullName AS PIN2,rel.chapter as CHAPTER
            ORDER BY rel.sequence
            '''
            query = query.format(label=label)
        else:
            raise ValueError
        data = Neo4j._graph.run(query).data()
        return data

    @staticmethod
    def get_chapters():
        query = '''
            MATCH (pin1:pin)-[rel]->(pin2:pin)
            RETURN DISTINCT rel.chapter as chapter
            ORDER BY chapter
            '''
        data = Neo4j._graph.run(query).data()
        return data

    def connector_status_dist(self):
        '''
        :return: json objects on pin tested times
        '''
        query = '''
        MATCH (pin1:pin)-[rel]-(pin2:pin)
        WHERE rel.status='PASS' OR rel.status='HIGH'
        RETURN count(rel.status) as Number,rel.status as Status,pin1.connectorName as Connector
        ORDER by Connector
        '''
        data = Neo4j._graph.run(query).data()
        if not data:
            data = [{'Number':0, 'Status':'NULL', 'Connector':None}]

        return data

    def connector_tested_pins(self):
        '''
        :return: json objects, tested pins number
        '''
        query = '''
        match (pin1:pin)-[rel]-(pin2:pin)
        WHERE rel.status='PASS' OR rel.status='HIGH'
        return DISTINCT pin1.fullName as pin ,pin1.connectorName as Connector
        Order by Connector
        '''
        data = Neo4j._graph.run(query).data()
        if not data:
            data = [{'Number':0, 'Status':'NULL', 'Connector':None}]

        return data




class FindFiles(object):
    def __init__(self, folder_in = None, fileExt = '.txt'):
        self._rawfiles = []
        try:
            for rt, dirs, files in os.walk(folder_in, topdown=False):
                for fl in files:

                    fl = os.path.join(rt, fl)
                    fl = os.path.abspath(fl)
                    f = os.path.splitext(fl)
                    if f[1] == fileExt:
                        self._rawfiles.append(fl)

        except:
            print("Error Happen! ")

    def paths(self):
        return self._rawfiles



class Format(object):
    def __init__(self, data):
        self._data = data

    def jsons_to_progTable(self,start=0):
        table = Table()
        table.columns = [u'No',u'测试程序',u'章节号',u'备注']
        table.array=[]
        r = 0
        for row in self._data:
            line1 = [r+start, u"X-"+row[u'PIN1'], '', '']
            line2 = ['', u"C-"+row[u'PIN2'], '', '']
            table.array.append(line1)
            table.array.append(line2)
            r += 1
        return table

    def jsons_to_DF(self):
        return pd.DataFrame(self._data)

    def jsons_to_testprog(self, start=0):
        '''
        '''
        col_data = [u"PIN1", u"PIN2", u"CHAPTER"]
        col_prog = [u"No", u"测试程序", u"章节号", u"备注"]
        No, pins, chapter = [], [], []
        high_line = self.jsons_DF()
        row, col = high_line.shape
        for i in range(row):
            No.append(str(start + i))
            pins.append(u"X-"+high_line[col_data[0]].iloc[i])
            chapter.append(high_line[col_data[2]].iloc[i])
            No.append("")
            pins.append(u"C-"+high_line[col_data[1]].iloc[i])
            chapter.append(high_line[col_data[2]].iloc[i])
        pd_prog = pd.DataFrame([], columns=col_prog)
        pd_prog[col_prog[0]] = No
        pd_prog[col_prog[1]] = pins
        pd_prog[col_prog[2]] = chapter

        return pd_prog

    # def to_txt_report(self, thr=0.5):
    #     '''
    #     Print out
    #     '''
    #     lst = self._stats_sort(threshold=thr)
    #     str_line = "===Node Name===PASS Ratio===\n"
    #     for item in lst:
    #         connector, passratio = item[0], item[1][2]
    #         str_line += "%12s%10.2f%%\n" % (connector, passratio * 100)
    #
    #     with open(self._report_out, 'w') as fp:
    #         fp.write(str_line)

class Save(object):
    keys = (u'chapter', u'pin1', u'pin2')
    col_name = (u'No', u'测试程序', u'章节号', u'备注')

    def __init__(self, data):
        self.Data = data
        pass

    def to_txt(self, path, fmt=None):
        column = self.pdData.columns
        n_row, n_col = self.pdData.shape
        header = ''
        for item in column:
            header += "{0:=^20}".format(item)
        header += "\n"
        lines = header
        for r in range(n_row):
            line = ''
            for item in self.pdData.iloc[r]:
                line += "{0:20}".format(item)
            line += '\n'
            lines += line
        with open(path, 'w') as fp:
            fp.write(lines)

    def to_csv(self, path, fmt=None):
        '''
        Print out
        '''
        self.pdData.to_csv(path, encoding='utf-8')

    def open_excel(self):
        ws = openpyxl.Workbook(write_only=True)
        return ws
 #       try:
  #          ws = openpyxl.Workbook(write_only=True)
   #         return ws
    #    except exception,e:
     #       print str(e)

    def to_excel2(self, fileName= 'file.xls'):
        table = Format(self.Data).jsons_to_progTable()
        ws = self.open_excel()
        sheet = ws.create_sheet()
        sheet.append(table.columns)
        for row in table.array:
            sheet.append(row)
        ws.save(fileName)

    def to_excel(self, path, sheet_name="sheet_1", fmt='noindex'):
        idx = True
        if fmt == "noindex":
            idx = False

        self.pdData.to_excel(path, sheet_name=sheet_name, encoding='utf-8',index=idx)

    def to_html(self, path, fmt=None):
        with open(path, 'wb') as fp:
            str_out = self.pdData.to_html()
            fp.write(str_out.encode('utf-8'))





def import_data():
    paths = FindFiles(folder_in="/Applications/Splunk/etc/apps/oats_admin/bin/import/", fileExt='.txt').paths()
    print(paths)
    db = Neo4j()
    sequence = 0
    for fp in paths:
        # tables = Jsw(fp)
        # sequence = db.jsw_upload(tables,offset=sequence)
        table = Pvg(fp)
        db.pvg_update(table)


def generate_prog(chapters=None,label=None):
    db = Neo4j()
    if not chapters:
        out = db.prog(chapter=None, label=label)
        Save(out).to_excel2(fileName='all_chapters.xlsx')
    else:
        for chp in chapters:
            out = db.prog(chapter=chp, label=label)
            Save(out).to_excel2(fileName=chp + '.xlsx')


if __name__ == "__main__":
    #import_data()
    generate_prog()
